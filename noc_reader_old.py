import os
import json
import logging
import asyncio
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Optional, Dict, Any, List

from dotenv import load_dotenv
from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openai import AsyncOpenAI
from slack_sdk.web.async_client import AsyncWebClient
from slack_sdk.signature import SignatureVerifier
import requests
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Import database and permissions modules
import db
import permissions

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('noc_reader_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Constants and configuration
UAE_TZ = ZoneInfo('Asia/Dubai')
EXCEL_FILE_PATH = 'noc_extractions.xlsx'
NOTIFIED_CACHE_PATH = 'noc_notified_cache.json'

# Validate required environment variables
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY')
if not OPENAI_API_KEY:
    logger.error('OPENAI_API_KEY not found in environment variables')
    raise ValueError('OPENAI_API_KEY is required')

# Slack configuration
SLACK_BOT_TOKEN = os.environ.get('SLACK_BOT_TOKEN')
SLACK_SIGNING_SECRET = os.environ.get('SLACK_SIGNING_SECRET')
if not SLACK_BOT_TOKEN:
    logger.error('SLACK_BOT_TOKEN not found in environment variables')
    raise ValueError('SLACK_BOT_TOKEN is required')
if not SLACK_SIGNING_SECRET:
    logger.error('SLACK_SIGNING_SECRET not found in environment variables')
    raise ValueError('SLACK_SIGNING_SECRET is required')

# Google Drive configuration
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', 'routes-key.json')
SHARED_DRIVE_ID = os.environ.get('SHARED_DRIVE_ID', '')
DRIVE_PARENT_ID = os.environ.get('DRIVE_PARENT_ID', '')  # optional folder ID within shared drive
DRIVE_FILE_ID = os.environ.get('DRIVE_FILE_ID', '')  # optional existing file id to update
SCOPES = ['https://www.googleapis.com/auth/drive']
drive_service = None
drive_file_id_cache: Optional[str] = None

# Initialize FastAPI app and OpenAI client
api = FastAPI(title='NOC Document Reader API', version='2.0')
client = AsyncOpenAI(api_key=OPENAI_API_KEY)
slack_client = AsyncWebClient(token=SLACK_BOT_TOKEN)
signature_verifier = SignatureVerifier(SLACK_SIGNING_SECRET)

# In-memory state
pending_confirmations: Dict[str, Dict[str, Any]] = {}
user_history: Dict[str, List[Dict[str, str]]]= {}

# ================ OpenAI Responses helpers ================
def _extract_responses_text(res: Any) -> str:
    """Safely extract textual content from a Responses API result across SDK variants."""
    try:
        # Prefer convenience
        text = getattr(res, 'output_text', None)
        if isinstance(text, str) and text.strip():
            return text
        # Older/newer shapes
        output = getattr(res, 'output', None)
        if isinstance(output, list) and len(output) > 0:
            first = output[0]
            content = getattr(first, 'content', None)
            if isinstance(content, list):
                # Scan from end to find last text-like part
                for part in reversed(content):
                    # SDK may expose objects or dicts
                    if hasattr(part, 'text') and getattr(part, 'text'):
                        return getattr(part, 'text')
                    if isinstance(part, dict) and part.get('text'):
                        return part['text']
        # Fallback: string repr
        return str(res)
    except Exception:
        return ""

# ================= Pydantic models =================
class NOCParseRequest(BaseModel):
    file_path: Optional[str] = None
    save_to_excel: bool = False
    submitted_by: str = 'API'

# ================= Excel utilities =================
async def initialize_excel() -> None:
    """Create Excel file with headers if it doesn't exist."""
    if not Path(EXCEL_FILE_PATH).exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'NOC Extractions'

        headers = [
            'ID',
            'Timestamp',
            'NOC Number',
            'Project Name',
            'Issue Date',
            'NOC Type',
            'Validity End Date',
            'Comments',
            'Submitted By',
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        await asyncio.to_thread(wb.save, EXCEL_FILE_PATH)

async def save_to_database(data: Dict[str, Any], document_path: Optional[str] = None) -> int:
    """Save parsed NOC data to database with optional document, return assigned ID."""
    try:
        document_data = None
        document_filename = None
        
        # Read document if provided
        if document_path and os.path.exists(document_path):
            try:
                with open(document_path, 'rb') as f:
                    document_data = f.read()
                document_filename = os.path.basename(document_path)
            except Exception as e:
                logger.error(f'Error reading document file: {e}')
        
        # Save to database
        record_id = await asyncio.to_thread(
            db.save_noc_extraction,
            noc_number=data.get('noc_number', ''),
            project_name=data.get('project_name', ''),
            issue_date=data.get('issue_date', ''),
            noc_type=data.get('noc_type', ''),
            validity_end_date=data.get('validity_end_date', ''),
            comments=data.get('comments', ''),
            submitted_by=data.get('submitted_by', ''),
            timestamp=datetime.now(UAE_TZ).isoformat(),
            raw_data=data,
            document_data=document_data,
            document_filename=document_filename
        )
        return record_id
    except Exception as e:
        logger.error(f'Error saving to database: {e}')
        return 0

# ================= Google Drive sync =================
def _ensure_drive_service():
    global drive_service
    if drive_service is not None:
        return
    try:
        creds = Credentials.from_service_account_file(GOOGLE_SERVICE_ACCOUNT_JSON, scopes=SCOPES)
        # Build Drive service
        globals()['drive_service'] = build('drive', 'v3', credentials=creds)
        logger.info('Initialized Google Drive service')
    except Exception as e:
        logger.warning(f'Google Drive service not initialized: {e}')

def _drive_find_file_by_name(name: str) -> Optional[str]:
    _ensure_drive_service()
    if drive_service is None:
        return None
    try:
        q_parts = [f"name = '{name}'", "trashed = false"]
        if DRIVE_PARENT_ID:
            q_parts.append(f"'{DRIVE_PARENT_ID}' in parents")
        query = ' and '.join(q_parts)
        resp = drive_service.files().list(
            q=query,
            corpora='drive' if SHARED_DRIVE_ID else 'user',
            driveId=SHARED_DRIVE_ID or None,
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            fields='files(id, name)',
            pageSize=1,
        ).execute()
        files = resp.get('files', [])
        if files:
            return files[0]['id']
    except Exception as e:
        logger.warning(f'Drive search failed: {e}')
    return None

def _drive_create_file(name: str, local_path: str) -> Optional[str]:
    _ensure_drive_service()
    if drive_service is None:
        return None
    try:
        metadata = {
            'name': name,
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        if DRIVE_PARENT_ID:
            metadata['parents'] = [DRIVE_PARENT_ID]
        media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=False)
        file = drive_service.files().create(
            body=metadata,
            media_body=media,
            fields='id, name',
            supportsAllDrives=True,
        ).execute()
        return file.get('id')
    except Exception as e:
        logger.warning(f'Drive create failed: {e}')
        return None

def _drive_update_file(file_id: str, local_path: str) -> bool:
    _ensure_drive_service()
    if drive_service is None:
        return False
    try:
        media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=False)
        drive_service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True,
        ).execute()
        return True
    except Exception as e:
        logger.warning(f'Drive update failed: {e}')
        return False

async def upload_excel_to_drive():
    """Upload or update the Excel file to Google Drive."""
    try:
        if not Path(EXCEL_FILE_PATH).exists():
            await initialize_excel()
        # Resolve file id
        global drive_file_id_cache
        file_id = DRIVE_FILE_ID or drive_file_id_cache or _drive_find_file_by_name(EXCEL_FILE_PATH)
        if not file_id:
            # create new file
            file_id = _drive_create_file(EXCEL_FILE_PATH, EXCEL_FILE_PATH)
            drive_file_id_cache = file_id
            if file_id:
                logger.info(f'Created Drive file for Excel: {file_id}')
        else:
            ok = _drive_update_file(file_id, EXCEL_FILE_PATH)
            if ok:
                drive_file_id_cache = file_id
                logger.info('Updated Excel on Google Drive')
    except Exception as e:
        logger.warning(f'Excel upload to Drive failed: {e}')

async def periodic_drive_upload(interval_seconds: int = 3600):
    while True:
        await upload_excel_to_drive()
        await asyncio.sleep(interval_seconds)

# ================= Expiry notifications =================
def _load_notified_cache() -> Dict[str, str]:
    try:
        if os.path.exists(NOTIFIED_CACHE_PATH):
            with open(NOTIFIED_CACHE_PATH, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_notified_cache(cache: Dict[str, str]) -> None:
    try:
        with open(NOTIFIED_CACHE_PATH, 'w') as f:
            json.dump(cache, f)
    except Exception:
        pass

def _parse_date_yyyy_mm_dd(value: str) -> Optional[date]:
    if not value or not isinstance(value, str):
        return None
    try:
        return datetime.strptime(value.strip(), '%Y-%m-%d').date()
    except Exception:
        # Try common alternatives
        for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d %b %Y', '%d %B %Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except Exception:
                continue
    return None

async def check_and_notify_expiring_nocs():
    """Notify Slack recipients if a NOC validity end date is within 14 days."""
    try:
        # Get expiring NOCs from database
        records = await asyncio.to_thread(db.get_expiring_nocs, 14)
    except Exception as e:
        logger.warning(f'Failed reading database for expiry checks: {e}')
        return

    recipients_env = os.environ.get('NOC_ALERT_RECIPIENTS', '')
    recipients = [r.strip() for r in recipients_env.split(',') if r.strip()]
    if not recipients:
        return

    cache = _load_notified_cache()
    today = datetime.now(UAE_TZ).date()

    for rec in records:
        noc_number = rec.get('noc_number') or ''
        project_name = rec.get('project_name') or ''
        end_str = rec.get('validity_end_date') or ''
        if not end_str:
            continue
        end_date = _parse_date_yyyy_mm_dd(str(end_str))
        if not end_date:
            continue
        days_left = (end_date - today).days
        if 0 < days_left <= 14:
            key = f"{noc_number}|{end_date.isoformat()}"
            last_notified = cache.get(key)
            if last_notified == today.isoformat():
                # already notified today
                continue
            msg = (
                f"Heads up: NOC `{noc_number}` for `{project_name or 'N/A'}` expires on `{end_date.isoformat()}` "
                f"(in {days_left} day{'s' if days_left != 1 else ''})."
            )
            for r in recipients:
                try:
                    await slack_client.chat_postMessage(channel=r, text=msg)
                except Exception as e:
                    logger.warning(f'Failed to send expiry alert to {r}: {e}')
            cache[key] = today.isoformat()
    _save_notified_cache(cache)

async def periodic_expiry_check(interval_seconds: int = 86400):
    while True:
        try:
            await check_and_notify_expiring_nocs()
        except Exception as e:
            logger.warning(f'Expiry check failed: {e}')
        await asyncio.sleep(interval_seconds)

async def migrate_existing_excel_to_db() -> int:
    """One-time migration from Excel to database."""
    if Path(EXCEL_FILE_PATH).exists():
        try:
            count = await asyncio.to_thread(db.migrate_from_excel, EXCEL_FILE_PATH)
            logger.info(f"Migrated {count} records from Excel to database")
            return count
        except Exception as e:
            logger.error(f"Error migrating Excel data: {e}")
    return 0

# ================= Parsing with OpenAI =================
async def parse_noc_with_ai(file_path: str) -> Dict[str, Any]:
    """Parse a NOC PDF using OpenAI Responses API and extract the key fields."""
    if not file_path:
        raise ValueError('file_path is required')
    if not file_path.lower().endswith('.pdf'):
        raise ValueError('Only PDF files are supported. Please provide a .pdf file path.')
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'File not found: {file_path}')

    uploaded_file_id: Optional[str] = None

    # Build system prompt with current date in UAE timezone
    now = datetime.now(UAE_TZ)
    today_str = now.strftime('%B %d, %Y')
    system_prompt = (
        f'You are an expert, helpful assistant that parses NOC (No Objection Certificate) documents.'
        f" Today's date is {today_str}.\n\n"
        'Your goal is to help the user by extracting accurate, normalized fields while keeping responses concise.'
        '\nExtract ONLY the following fields from the attached NOC PDF and return a STRICT JSON object:'
        '\n- noc_number: The NOC reference number or ID (choose the most prominent/official if multiple are present)'
        '\n- project_name: The official project name/title associated with the NOC (if multiple, pick the one tied to the NOC reference)'
        '\n- issue_date: The date the NOC was issued (prefer issuance over expiry/renewal; format as YYYY-MM-DD when possible)'
        '\n- noc_type: The NOC type (e.g., company formation, activity change, etc.; keep it short and specific)'
        '\n- validity_end_date: The validity end date (date to/expiry), if present (format as YYYY-MM-DD when possible)'
        "\n- comments: Brief human remarks or notes present in the document. EXCLUDE procedural instructions (e.g., 'bring original', 'pay fee', 'submit within 7 days')."
        '\nIf any field is not present, return an empty string for it.'
        '\nDo not include any commentary outside of the JSON output.'
        '\nNormalize values: trim labels/punctuation, and standardize dates to YYYY-MM-DD when possible.'
    )

    try:
        # Upload the PDF file
        with open(file_path, 'rb') as f:
            upload = await client.files.create(file=f, purpose='user_data')
        uploaded_file_id = upload.id
        logger.info(f'Uploaded file to OpenAI: {uploaded_file_id}')

        content = [
            {"type": "input_file", "file_id": uploaded_file_id},
            {
                "type": "input_text",
                "text": (
                    'Please read the PDF and output a JSON object matching the provided schema.'
                ),
            },
        ]

        # Call Responses API with enforced JSON schema
        res = await client.responses.create(
            model='gpt-5',
            input=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": content},
            ],
            text={
                'format': {
                    'type': 'json_schema',
                    'name': 'noc_extraction_schema',
                    'strict': True,
                    'schema': {
                        'type': 'object',
                        'properties': {
                            'noc_number': {'type': 'string'},
                            'project_name': {'type': 'string'},
                            'issue_date': {'type': 'string', 'description': 'YYYY-MM-DD when possible'},
                            'noc_type': {'type': 'string'},
                            'validity_end_date': {'type': 'string', 'description': 'YYYY-MM-DD when possible'},
                            'comments': {'type': 'string'},
                        },
                        'required': ['noc_number', 'project_name', 'issue_date', 'noc_type', 'validity_end_date', 'comments'],
                        'additionalProperties': False,
                    },
                },
            },
            store=False,
        )

        # Best-effort JSON extraction from Responses output
        parsed: Dict[str, Any] = {}
        try:
            text_value = _extract_responses_text(res)
            if text_value:
                parsed = json.loads(text_value)
        except Exception as parse_err:
            logger.warning(f'Fallback parsing of response failed: {parse_err}')
            raise

        # Attach timestamp for provenance
        parsed['timestamp'] = datetime.now(UAE_TZ).isoformat()
        return parsed

    except Exception as e:
        logger.error(f'Error parsing NOC with AI: {e}')
        return {}
    finally:
        # Cleanup uploaded file
        if uploaded_file_id:
            try:
                await client.files.delete(uploaded_file_id)
            except Exception as cleanup_err:
                logger.warning(f'Failed to delete uploaded file: {cleanup_err}')

# ================= Conversation helpers =================
def append_to_history(user_id: str, role: str, content: str) -> None:
    if user_id not in user_history:
        user_history[user_id] = []
    user_history[user_id].append({"role": role, "content": content})
    if len(user_history[user_id]) > 20:
        user_history[user_id] = user_history[user_id][-20:]

def render_summary(noc_data: Dict[str, Any]) -> str:
    return (
        f"• NOC Number: `{noc_data.get('noc_number','') or 'N/A'}`\n"
        f"• Project Name: `{noc_data.get('project_name','') or 'N/A'}`\n"
        f"• Issue Date: `{noc_data.get('issue_date','') or 'N/A'}`\n"
        f"• NOC Type: `{noc_data.get('noc_type','') or 'N/A'}`\n"
        f"• Validity End Date: `{noc_data.get('validity_end_date','') or 'N/A'}`\n"
        (f"• Comments: {noc_data.get('comments','')[:300]}\n" if noc_data.get('comments') else '')
    )

def format_noc_results(noc_data: Dict[str, Any]) -> str:
    return (
        "**NOC Extraction Result**\n\n" + render_summary(noc_data) +
        "Reply 'yes' to confirm saving, or provide corrections (e.g., 'Change date to 2024-05-01' or 'NOC number is 123'), or say 'cancel'."
    )

async def llm_confirmation_decision(user_id: str, user_text: str, pending: Dict[str, Any]) -> Dict[str, Any]:
    """Use LLM to decide whether the user is confirming, editing fields, or cancelling."""
    today_str = datetime.now(UAE_TZ).strftime('%B %d, %Y')
    system_prompt = (
        "You are a friendly, concise Slack assistant helping confirm or edit extracted NOC fields. "
        f"Today's date is {today_str}. "
        "Return STRICT JSON describing the action to take. "
        "Allowed actions: 'confirm', 'cancel', 'edit', 'clarify'. "
        "If the user confirms, set action to 'confirm'. "
        "If the user cancels, set action to 'cancel'. "
        "If the user provides corrections, set action to 'edit' and include only corrected fields in 'fields'. "
        "If you need more info, set action to 'clarify'. "
        "Always craft a short, friendly Slack-style 'message' (<= 2 sentences) that acknowledges the user's intent and guides the next step. "
        "When asking to confirm, include a brief Slack-formatted summary of the current fields inline in 'message' (use backticks for values). "
        "When action is 'edit', briefly summarize applied changes in 'message'. "
        "When action is 'clarify', ask exactly one specific question."
    )

    pending_text = (
        "Pending NOC fields to confirm:\n" + render_summary(pending) +
        "User message: " + (user_text or '')
    )

    # Build conversation with short recent history
    history = user_history.get(user_id, [])[-8:]
    messages = [{"role": "system", "content": system_prompt}] + history + [
        {"role": "user", "content": pending_text}
    ]

    res = await client.responses.create(
        model='gpt-5',
        input=messages,
        text={
            'format': {
                'type': 'json_schema',
                'name': 'noc_confirmation_schema',
                'strict': False,
                'schema': {
                    'type': 'object',
                    'properties': {
                        'action': {'type': 'string', 'enum': ['confirm', 'cancel', 'edit', 'clarify']},
                        'fields': {
                            'type': 'object',
                            'properties': {
                                'noc_number': {'type': 'string'},
                                'project_name': {'type': 'string'},
                                'issue_date': {'type': 'string'},
                                'noc_type': {'type': 'string'},
                                'validity_end_date': {'type': 'string'},
                                'comments': {'type': 'string'},
                            },
                            'additionalProperties': False
                        },
                        'message': {'type': 'string'}
                    },
                    'required': ['action'],
                    'additionalProperties': False
                },
            }
        },
        store=False
    )

    # Parse JSON result
    parsed: Dict[str, Any] = {}
    try:
        text_value = _extract_responses_text(res)
        if text_value:
            parsed = json.loads(text_value)
    except Exception as e:
        logger.error(f'Error parsing confirmation response: {e}')
        parsed = {'action': 'clarify', 'message': "I couldn't understand. Please reply 'yes' to save, or provide corrected fields, or say 'cancel'."}

    return parsed

# ================= Slack helpers =================
async def download_slack_file(file_info: Dict[str, Any]) -> str:
    """Download a Slack file locally and return the path."""
    try:
        file_id = file_info.get('id')
        file_name = file_info.get('name', 'document.pdf')
        file_response = await slack_client.files_info(file=file_id)
        if not file_response.get('ok'):
            raise RuntimeError(f'Failed to get file info: {file_response}')

        file_data = file_response['file']
        file_url = file_data.get('url_private_download') or file_data.get('url_private')
        if not file_url:
            raise RuntimeError('No downloadable URL for file')

        os.makedirs('./uploads', exist_ok=True)
        safe_name = ''.join(ch if ch.isalnum() or ch in ('-', '_', '.') else '_' for ch in file_name)
        timestamp = datetime.now(UAE_TZ).strftime('%Y%m%d_%H%M%S')
        local_path = f'./uploads/noc_{timestamp}_{safe_name}'

        headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}
        resp = requests.get(file_url, headers=headers)
        resp.raise_for_status()
        with open(local_path, 'wb') as f_out:
            f_out.write(resp.content)
        logger.info(f'Downloaded Slack file to {local_path}')
        return local_path
    except Exception as e:
        logger.error(f'Error downloading Slack file: {e}')
        return ''

async def handle_slack_message(channel: str, user_id: str, text: str, files: Optional[List[Dict[str, Any]]]) -> None:
    """Handle Slack message: manage confirmation flow and file parsing via LLM."""
    # Resolve user display name
    try:
        user_info = await slack_client.users_info(user=user_id)
        profile = user_info["user"]["profile"]
        display_name = profile.get("display_name") or profile.get("real_name") or "Unknown"
    except Exception:
        display_name = "Unknown"

    # Append user message to history
    if text:
        append_to_history(user_id, 'user', text)

    # Confirmation loop via LLM
    if user_id in pending_confirmations:
        pending = pending_confirmations[user_id]
        decision = await llm_confirmation_decision(user_id, text or '', pending)
        action = decision.get('action')
        message = decision.get('message')

        if action == 'confirm':
            pending['submitted_by'] = display_name
            document_path = pending.pop('_document_path', None)
            record_id = await save_to_database(pending, document_path)
            
            # Clean up document file after saving
            if document_path:
                try:
                    os.remove(document_path)
                except Exception:
                    pass
            
            if record_id:
                reply = f"Saved. Record ID: #{record_id}. NOC: `{pending.get('noc_number','')}`"
            else:
                reply = 'Failed to save NOC information. Please try again.'
            pending_confirmations.pop(user_id, None)
            append_to_history(user_id, 'assistant', reply)
            await slack_client.chat_postMessage(channel=channel, text=reply)
            return
        elif action == 'cancel':
            # Clean up document file on cancel
            document_path = pending.pop('_document_path', None)
            if document_path:
                try:
                    os.remove(document_path)
                except Exception:
                    pass
            
            pending_confirmations.pop(user_id, None)
            reply = message or 'Cancelled. No data was saved.'
            append_to_history(user_id, 'assistant', reply)
            await slack_client.chat_postMessage(channel=channel, text=reply)
            return
        elif action == 'edit':
            fields = decision.get('fields', {}) or {}
            for k in ('noc_number', 'issue_date', 'noc_type'):
                if k in fields and isinstance(fields[k], str) and fields[k].strip() != '':
                    pending[k] = fields[k].strip()
            # Keep pending, ask for confirm again (let LLM craft the message)
            reply = message or "Please review the updated fields and confirm."
            append_to_history(user_id, 'assistant', reply)
            await slack_client.chat_postMessage(channel=channel, text=reply)
            return
        else:  # clarify or unknown
            reply = message or "Please reply 'yes' to save, provide corrections, or say 'cancel'."
            append_to_history(user_id, 'assistant', reply)
            await slack_client.chat_postMessage(channel=channel, text=reply)
            return

    # If PDF file is attached, process it
    if files:
        pdf_file = next((f for f in files if (f.get('mimetype') == 'application/pdf' or f.get('name','').lower().endswith('.pdf'))), None)
        if pdf_file:
            local_path = await download_slack_file(pdf_file)
            if local_path and local_path.lower().endswith('.pdf'):
                noc_data = await parse_noc_with_ai(local_path)
                
                if noc_data:
                    # Store file path for later saving
                    pending_confirmations[user_id] = noc_data
                    pending_confirmations[user_id]['_document_path'] = local_path
                    # Ask LLM to craft the initial confirmation ask, including a brief summary
                    decision = await llm_confirmation_decision(user_id, '', noc_data)
                    reply = decision.get('message') or "I've extracted the fields. Please review and confirm."
                    append_to_history(user_id, 'assistant', reply)
                    await slack_client.chat_postMessage(channel=channel, text=reply)
                else:
                    # Clean up file on failure
                    try:
                        os.remove(local_path)
                    except Exception:
                        pass
                    fail_msg = 'Failed to extract NOC information from the PDF.'
                    append_to_history(user_id, 'assistant', fail_msg)
                    await slack_client.chat_postMessage(channel=channel, text=fail_msg)
                return

    # Use LLM with tools to handle all text-based requests
    is_admin = permissions.is_admin(user_id)
    reply = await llm_general_response(user_id, text or '', channel, is_admin)
    append_to_history(user_id, 'assistant', reply)
    await slack_client.chat_postMessage(channel=channel, text=reply)

async def llm_general_response(user_id: str, user_text: str, channel: str, is_admin: bool = False) -> str:
    """Use LLM with tools to handle various user requests."""
    today_str = datetime.now(UAE_TZ).strftime('%B %d, %Y')
    
    admin_note = ""
    if is_admin:
        admin_note = """
        Admin features available:
        - Export database to Excel
        - Retrieve specific NOC documents by reference number
        - View detailed statistics
        """
    
    system_prompt = f"""
You are a friendly Slack assistant that helps with NOC (No Objection Certificate) document processing.
Today's date is {today_str}.

FEATURES:
- Process NOC PDFs: Extract details from uploaded documents
- Database Summary: Show statistics and recent NOCs
- Export Database: Export all records to Excel (admin only)
- Retrieve Documents: Get specific NOC documents by reference number (admin only)

{admin_note}

IMPORTANT:
- Keep responses brief and friendly
- Use Slack formatting (backticks, bullets)
- When user mentions a NOC reference number with intent to retrieve, use get_noc_document tool
- When user wants to export/download the database, use export_database tool
- When user wants statistics/summary, use get_summary tool
"""
    
    history = user_history.get(user_id, [])[-8:]
    messages = ([{"role": "system", "content": system_prompt}] + history +
                [{"role": "user", "content": user_text or ' '}])
    
    tools = []
    
    # Add admin tools if user is admin
    if is_admin:
        tools.extend([
            {
                "type": "function",
                "name": "get_noc_document",
                "description": "Retrieve a specific NOC document by reference number",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "noc_number": {
                            "type": "string",
                            "description": "The NOC reference number to retrieve"
                        }
                    },
                    "required": ["noc_number"]
                }
            },
            {
                "type": "function",
                "name": "export_database",
                "description": "Export the entire NOC database to Excel file",
                "parameters": {
                    "type": "object",
                    "properties": {}
                }
            }
        ])
    
    # Add general tools
    tools.append({
        "type": "function",
        "name": "get_summary",
        "description": "Get database summary with statistics",
        "parameters": {
            "type": "object",
            "properties": {}
        }
    })
    
    try:
        res = await client.responses.create(
            model='gpt-5',
            input=messages,
            tools=tools if tools else None,
            tool_choice="auto" if tools else None,
            store=False
        )
        
        # Check if LLM wants to use a tool
        if res.output and len(res.output) > 0:
            msg = res.output[0]
            
            if hasattr(msg, 'type') and msg.type == "function_call":
                if msg.name == "get_noc_document" and is_admin:
                    args = json.loads(msg.arguments)
                    noc_number = args.get("noc_number", "").strip()
                    
                    # Get document
                    doc_result = await asyncio.to_thread(db.get_noc_document, noc_number)
                    
                    if doc_result:
                        document_data, filename = doc_result
                        
                        # Get NOC details
                        noc_details = await asyncio.to_thread(db.get_noc_by_number, noc_number)
                        
                        # Create temp file and upload
                        import tempfile
                        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                        temp_file.write(document_data)
                        temp_file.close()
                        
                        # Upload to Slack
                        await slack_client.files_upload_v2(
                            channel=channel,
                            file=open(temp_file.name, 'rb'),
                            filename=filename,
                            title=f'NOC {noc_number}',
                            initial_comment=f"**NOC Details:**\n• Project: {noc_details.get('project_name', 'N/A') if noc_details else 'N/A'}\n• Type: {noc_details.get('noc_type', 'N/A') if noc_details else 'N/A'}\n• Issue Date: {noc_details.get('issue_date', 'N/A') if noc_details else 'N/A'}\n• Validity End Date: {noc_details.get('validity_end_date', 'N/A') if noc_details else 'N/A'}"
                        )
                        
                        os.unlink(temp_file.name)
                        return f"✅ Found and uploaded NOC `{noc_number}`"
                    else:
                        return f"❌ Could not find document for NOC `{noc_number}`. It may not have a stored document or doesn't exist."
                
                elif msg.name == "export_database" and is_admin:
                    # Export database
                    excel_path = await asyncio.to_thread(db.export_to_excel)
                    
                    # Upload to Slack
                    with open(excel_path, 'rb') as f:
                        await slack_client.files_upload_v2(
                            channel=channel,
                            file=f,
                            filename=f'noc_export_{datetime.now(UAE_TZ).strftime("%Y%m%d_%H%M%S")}.xlsx',
                            title='NOC Database Export',
                            initial_comment='Here is the complete NOC database export.'
                        )
                    
                    os.unlink(excel_path)
                    
                    total = len(await asyncio.to_thread(db.get_all_noc_extractions))
                    return f"✅ Database exported successfully. Total records: {total}"
                
                elif msg.name == "get_summary":
                    summary = await asyncio.to_thread(db.get_noc_summary)
                    
                    reply = f"📊 *NOC Database Summary*\n\n"
                    reply += f"Total NOCs: *{summary['total_nocs']}*\n"
                    reply += f"Archived NOCs: *{summary.get('history_nocs', 0)}*\n\n"
                    
                    if summary['by_type']:
                        reply += "*By Type:*\n"
                        for noc_type, count in summary['by_type'].items():
                            reply += f"• {noc_type}: {count}\n"
                        reply += "\n"
                    
                    if summary['expiring_soon']:
                        reply += f"*Expiring Soon ({len(summary['expiring_soon'])} NOCs):*\n"
                        for noc in summary['expiring_soon'][:3]:
                            reply += f"• `{noc['noc_number']}` - {noc['project_name'] or 'N/A'} (expires: {noc['validity_end_date']})\n"
                        if len(summary['expiring_soon']) > 3:
                            reply += f"_...and {len(summary['expiring_soon']) - 3} more_\n"
                        reply += "\n"
                    
                    if summary['recent_nocs']:
                        reply += "*Recent NOCs:*\n"
                        for noc in summary['recent_nocs'][:3]:
                            reply += f"• `{noc['noc_number']}` - {noc['project_name'] or 'N/A'}\n"
                    
                    return reply
            
            # Extract regular text response
            text_value = _extract_responses_text(res)
            if text_value:
                return text_value
        
    except Exception as e:
        logger.warning(f'LLM response failed: {e}')
    
    return 'Hi! You can upload a NOC PDF and I will extract the NOC number, issue date, and type.'

# ================= FastAPI endpoints =================
@api.post('/cron/check-expiry')
async def cron_check_expiry(request: Request):
    """Cron endpoint to check for expiring and expired NOCs."""
    try:
        # Verify cron secret if provided
        cron_secret = os.environ.get('CRON_SECRET')
        if cron_secret:
            auth_header = request.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer ') or auth_header[7:] != cron_secret:
                return JSONResponse({'error': 'Unauthorized'}, status_code=401)
        
        # Archive expired NOCs
        archived = await asyncio.to_thread(db.archive_expired_nocs)
        
        # Notify admins about archived NOCs
        if archived:
            admin_channels = permissions.get_admin_notification_channels()
            if not admin_channels:
                logger.warning("[CRON] No admin channels configured for notifications")
            else:
                for noc_number in archived:
                    # Get NOC details before it was archived
                    msg = f"📦 *NOC Archived*\nNOC `{noc_number}` has expired and been moved to history."
                    
                    # Try to get document and send it
                    doc_result = await asyncio.to_thread(db.get_noc_document, noc_number)
                    
                    notification_sent = False
                    for channel in admin_channels:
                        try:
                            if doc_result:
                                document_data, filename = doc_result
                                # Create temp file
                                import tempfile
                                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                                temp_file.write(document_data)
                                temp_file.close()
                                
                                # Upload to Slack
                                with open(temp_file.name, 'rb') as f:
                                    await slack_client.files_upload_v2(
                                        channel=channel,
                                        file=f,
                                        filename=filename,
                                        title=f'Archived NOC {noc_number}',
                                        initial_comment=msg
                                    )
                                os.unlink(temp_file.name)
                            else:
                                await slack_client.chat_postMessage(channel=channel, text=msg)
                            notification_sent = True
                        except Exception as e:
                            logger.error(f'Failed to send archive notification to {channel}: {e}')
                    
                    if not notification_sent:
                        logger.error(f"Failed to send archive notification for NOC {noc_number} to any channel")
        
        # Check for NOCs expiring in 14 days
        expiring = await asyncio.to_thread(db.get_expiring_nocs, 14)
        notified = []
        
        today_str = datetime.now().date().isoformat()
        admin_channels = permissions.get_admin_notification_channels()
        
        if not admin_channels:
            logger.warning("[CRON] No admin channels configured for expiry notifications")
        
        for noc in expiring:
            noc_number = noc.get('noc_number', '')
            last_notified = noc.get('last_notified_date', '')
            
            # Only notify once per day
            if last_notified == today_str:
                continue
            
            project_name = noc.get('project_name', 'N/A')
            validity_end = noc.get('validity_end_date', '')
            days_left = (datetime.strptime(validity_end, '%Y-%m-%d').date() - datetime.now().date()).days
            
            msg = (
                f"⚠️ *NOC Expiring Soon*\n"
                f"NOC `{noc_number}` for project `{project_name}` expires in *{days_left} days* "
                f"(on {validity_end})."
            )
            
            # Get document
            doc_result = await asyncio.to_thread(db.get_noc_document, noc_number)
            
            # Send to all admin channels
            notification_sent = False
            for channel in admin_channels:
                try:
                    if doc_result:
                        document_data, filename = doc_result
                        # Create temp file
                        import tempfile
                        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                        temp_file.write(document_data)
                        temp_file.close()
                        
                        # Upload to Slack
                        with open(temp_file.name, 'rb') as f:
                            await slack_client.files_upload_v2(
                                channel=channel,
                                file=f,
                                filename=filename,
                                title=f'NOC {noc_number} - Expiring Soon',
                                initial_comment=msg
                            )
                        os.unlink(temp_file.name)
                    else:
                        await slack_client.chat_postMessage(channel=channel, text=msg)
                    notification_sent = True
                except Exception as e:
                    logger.error(f'Failed to send expiry notification to {channel}: {e}')
            
            # Update last notified date only if we sent at least one notification
            if notification_sent:
                await asyncio.to_thread(db.update_last_notified, noc_number, today_str)
                notified.append(noc_number)
            else:
                logger.error(f"Failed to send expiry notification for NOC {noc_number} to any channel")
        
        return JSONResponse({
            'success': True,
            'archived': len(archived),
            'notified_expiring': len(notified),
            'timestamp': datetime.now(UAE_TZ).isoformat()
        })
        
    except Exception as e:
        logger.error(f'Cron check-expiry error: {e}')
        return JSONResponse({'success': False, 'error': str(e)}, status_code=500)


@api.post('/api/parse_noc')
async def api_parse_noc(request: NOCParseRequest):
    """Parse a NOC PDF and optionally save results to Excel."""
    try:
        if not request.file_path:
            return JSONResponse({
                'success': False,
                'message': 'No file path provided',
            }, status_code=400)

        noc_data = await parse_noc_with_ai(request.file_path)
        if not noc_data:
            return JSONResponse({
                'success': False,
                'message': 'Failed to parse NOC document',
            }, status_code=500)

        if request.save_to_excel:
            noc_data['submitted_by'] = request.submitted_by
            record_id = await save_to_database(noc_data)
            noc_data['record_id'] = record_id

        return JSONResponse({'success': True, 'data': noc_data})
    except Exception as e:
        logger.exception('Unhandled error in api_parse_noc')
        return JSONResponse({'success': False, 'error': str(e)}, status_code=500)

@api.get('/api/get_nocs')
async def api_get_nocs():
    """Retrieve all stored NOC extractions."""
    try:
        records = await asyncio.to_thread(db.get_all_noc_extractions)
        return JSONResponse({'success': True, 'records': records, 'count': len(records)})
    except Exception as e:
        return JSONResponse({'success': False, 'error': str(e)}, status_code=500)

@api.get('/api/health')
async def health_check():
    return {'status': 'healthy', 'timestamp': datetime.now(UAE_TZ).isoformat()}

# ================= Slack endpoints =================
@api.post('/slack/events')
async def slack_events(request: Request):
    """Slack events endpoint for messages and file uploads."""
    body_bytes = await request.body()
    timestamp = request.headers.get('X-Slack-Request-Timestamp', '')
    signature = request.headers.get('X-Slack-Signature', '')

    if not signature_verifier.is_valid(body_bytes.decode(), timestamp, signature):
        return JSONResponse({'error': 'Invalid signature'}, status_code=403)

    data = await request.json()
    if data.get('type') == 'url_verification':
        return JSONResponse({'challenge': data.get('challenge')})

    if data.get('type') == 'event_callback':
        event = data.get('event', {})
        if event.get('type') == 'message' and not event.get('bot_id'):
            user_id = event.get('user')
            channel = event.get('channel')
            text = event.get('text', '')
            files = event.get('files', [])
            asyncio.create_task(handle_slack_message(channel, user_id, text, files))
            return JSONResponse({'status': 'ok'})

    return JSONResponse({'status': 'ok'})

# ================= Entrypoint =================
async def main():
    # Migrate existing Excel data to database
    await migrate_existing_excel_to_db()

    # Run FastAPI via uvicorn
    import uvicorn
    config = uvicorn.Config(app=api, host='0.0.0.0', port=3000, log_level='info')
    server = uvicorn.Server(config)

    logger.info('NOC Reader API running on http://localhost:3000')
    logger.info('API docs available at http://localhost:3000/docs')
    logger.info('Slack events endpoint: http://localhost:3000/slack/events')

    await server.serve()

if __name__ == '__main__':
    asyncio.run(main())